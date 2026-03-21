"""Import API endpoints."""

import uuid
import csv
import io
from datetime import datetime

from fastapi import APIRouter, Depends, Query, status, UploadFile, File
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text

from app.api.deps import CurrentUser, get_session
from app.core.exceptions import ForbiddenException
from app.core.permissions import is_admin_or_above

router = APIRouter()


class ImportBatchResponse(BaseModel):
    id: int
    batch_no: str
    import_type: str
    file_name: str
    status: str
    total_rows: int
    success_rows: int
    failed_rows: int
    started_at: str
    finished_at: str | None
    created_at: str


class ImportBatchListResponse(BaseModel):
    items: list[ImportBatchResponse]
    total: int
    page: int
    page_size: int


class ImportErrorResponse(BaseModel):
    id: int
    row_no: int
    error_code: str
    error_message: str
    row_data: dict | None


class ImportErrorListResponse(BaseModel):
    items: list[ImportErrorResponse]
    total: int


@router.get("/batches", response_model=ImportBatchListResponse)
async def list_import_batches(
    current_user: CurrentUser,
    session: AsyncSession = Depends(get_session),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    import_type: str | None = Query(None),
):
    """List import batches."""
    if not is_admin_or_above(current_user.role_name):
        raise ForbiddenException("Admin access required")

    offset = (page - 1) * page_size

    base_where = "1=1"
    params = {}

    if import_type:
        base_where += " AND import_type = :import_type"
        params["import_type"] = import_type

    # Count
    count_query = f"SELECT COUNT(*) FROM import_batches WHERE {base_where}"
    result = await session.execute(count_query, params)
    total = result.scalar()

    # Data
    data_query = f"""
        SELECT id, batch_no, import_type, file_name, status, total_rows, success_rows, failed_rows, started_at, finished_at, created_at
        FROM import_batches
        WHERE {base_where}
        ORDER BY created_at DESC
        OFFSET :offset LIMIT :limit
    """
    params["offset"] = offset
    params["limit"] = page_size

    result = await session.execute(data_query, params)
    rows = result.fetchall()

    items = [
        ImportBatchResponse(
            id=row[0],
            batch_no=row[1],
            import_type=row[2],
            file_name=row[3],
            status=row[4],
            total_rows=row[5],
            success_rows=row[6],
            failed_rows=row[7],
            started_at=row[8].isoformat() if row[8] else "",
            finished_at=row[9].isoformat() if row[9] else None,
            created_at=row[10].isoformat() if row[10] else "",
        )
        for row in rows
    ]

    return ImportBatchListResponse(items=items, total=total, page=page, page_size=page_size)


@router.get("/batches/{batch_id}/errors", response_model=ImportErrorListResponse)
async def list_import_errors(
    batch_id: int,
    current_user: CurrentUser,
    session: AsyncSession = Depends(get_session),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
):
    """List import errors for a batch."""
    if not is_admin_or_above(current_user.role_name):
        raise ForbiddenException("Admin access required")

    offset = (page - 1) * page_size

    # Count
    count_query = "SELECT COUNT(*) FROM import_errors WHERE batch_id = :batch_id"
    result = await session.execute(count_query, {"batch_id": batch_id})
    total = result.scalar()

    # Data
    data_query = """
        SELECT id, row_no, error_code, error_message, row_data
        FROM import_errors
        WHERE batch_id = :batch_id
        ORDER BY row_no
        OFFSET :offset LIMIT :limit
    """
    result = await session.execute(data_query, {"batch_id": batch_id, "offset": offset, "limit": page_size})
    rows = result.fetchall()

    items = [
        ImportErrorResponse(
            id=row[0],
            row_no=row[1],
            error_code=row[2],
            error_message=row[3],
            row_data=row[4] if row[4] else None,
        )
        for row in rows
    ]

    return ImportErrorListResponse(items=items, total=total)


async def _parse_excel_or_csv(file: UploadFile) -> list[dict]:
    """Parse Excel or CSV file and return list of rows as dicts."""
    content = await file.read()
    filename = file.filename or ""

    if filename.endswith('.csv'):
        # Parse CSV
        text_content = content.decode('utf-8-sig')  # Handle BOM
        reader = csv.DictReader(io.StringIO(text_content))
        return list(reader)
    elif filename.endswith(('.xlsx', '.xls')):
        # Parse Excel using openpyxl
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        if ws is None:
            return []

        # Get header row
        headers = [cell.value for cell in ws[1] if cell.value]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):  # Skip empty rows
                row_dict = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        row_dict[header] = row[i]
                rows.append(row_dict)
        return rows
    else:
        raise ValueError("Unsupported file format. Please use .xlsx, .xls, or .csv")


async def _find_customer_by_phone(session: AsyncSession, phone: str) -> int | None:
    """Find customer ID by phone number."""
    query = text("SELECT id FROM customers WHERE phone = :phone AND deleted_at IS NULL")
    result = await session.execute(query, {"phone": phone})
    row = result.fetchone()
    return row[0] if row else None


async def _record_import_error(
    session: AsyncSession,
    batch_id: int,
    row_no: int,
    error_code: str,
    error_message: str,
    row_data: dict
):
    """Record an import error."""
    query = text("""
        INSERT INTO import_errors (batch_id, row_no, error_code, error_message, row_data, created_at)
        VALUES (:batch_id, :row_no, :error_code, :error_message, :row_data, NOW())
    """)
    await session.execute(query, {
        "batch_id": batch_id,
        "row_no": row_no,
        "error_code": error_code,
        "error_message": error_message,
        "row_data": str(row_data),
    })


@router.post("/course-purchases", response_model=ImportBatchResponse, status_code=status.HTTP_201_CREATED)
async def import_course_purchases(
    current_user: CurrentUser,
    session: AsyncSession = Depends(get_session),
    file: UploadFile = File(...),
):
    """Import course purchase records from Excel/CSV.

    Required columns: customer_phone, course_name, purchase_date, amount
    Optional columns: currency
    """
    if not is_admin_or_above(current_user.role_name):
        raise ForbiddenException("Admin access required")

    # Generate batch number
    batch_no = f"CP{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:6].upper()}"

    # Create batch record
    insert_query = text("""
        INSERT INTO import_batches (batch_no, import_type, file_name, status, total_rows, success_rows, failed_rows, started_at, created_by, created_at, updated_at)
        VALUES (:batch_no, 'course_purchase', :file_name, 'processing', 0, 0, 0, NOW(), :created_by, NOW(), NOW())
        RETURNING id
    """)
    result = await session.execute(insert_query, {
        "batch_no": batch_no,
        "file_name": file.filename,
        "created_by": current_user.id,
    })
    batch_id = result.fetchone()[0]

    try:
        # Parse file
        rows = await _parse_excel_or_csv(file)
        total_rows = len(rows)
        success_rows = 0
        failed_rows = 0

        for idx, row_data in enumerate(rows, start=2):  # Start from 2 (1 is header)
            try:
                # Validate required fields
                phone = str(row_data.get('customer_phone', '')).strip()
                course_name = str(row_data.get('course_name', '')).strip()
                purchase_date = row_data.get('purchase_date')
                amount = row_data.get('amount')

                if not phone:
                    await _record_import_error(session, batch_id, idx, "MISSING_PHONE", "缺少客户手机号", dict(row_data))
                    failed_rows += 1
                    continue

                if not course_name:
                    await _record_import_error(session, batch_id, idx, "MISSING_COURSE", "缺少课程名称", dict(row_data))
                    failed_rows += 1
                    continue

                if not purchase_date:
                    await _record_import_error(session, batch_id, idx, "MISSING_DATE", "缺少购买日期", dict(row_data))
                    failed_rows += 1
                    continue

                if amount is None:
                    await _record_import_error(session, batch_id, idx, "MISSING_AMOUNT", "缺少金额", dict(row_data))
                    failed_rows += 1
                    continue

                # Find customer
                customer_id = await _find_customer_by_phone(session, phone)
                if not customer_id:
                    await _record_import_error(session, batch_id, idx, "CUSTOMER_NOT_FOUND", f"找不到手机号为 {phone} 的客户", dict(row_data))
                    failed_rows += 1
                    continue

                # Parse date
                if isinstance(purchase_date, str):
                    from dateutil import parser
                    purchase_date = parser.parse(purchase_date).date()
                elif hasattr(purchase_date, 'date'):
                    purchase_date = purchase_date.date()

                # Parse amount
                amount = float(amount)

                # Insert record
                insert_record = text("""
                    INSERT INTO course_purchase_records
                    (customer_id, course_name, purchase_date, amount, currency, import_batch_id, import_source, created_at, updated_at)
                    VALUES (:customer_id, :course_name, :purchase_date, :amount, :currency, :batch_id, 'excel', NOW(), NOW())
                """)
                await session.execute(insert_record, {
                    "customer_id": customer_id,
                    "course_name": course_name,
                    "purchase_date": purchase_date,
                    "amount": amount,
                    "currency": str(row_data.get('currency', 'CNY')),
                    "batch_id": batch_id,
                })
                success_rows += 1

            except Exception as e:
                await _record_import_error(session, batch_id, idx, "PARSE_ERROR", str(e), dict(row_data))
                failed_rows += 1

        # Update batch status
        status_val = "completed" if failed_rows == 0 else ("partial_success" if success_rows > 0 else "failed")
        update_query = text("""
            UPDATE import_batches
            SET status = :status, total_rows = :total, success_rows = :success, failed_rows = :failed, finished_at = NOW(), updated_at = NOW()
            WHERE id = :id
        """)
        await session.execute(update_query, {
            "status": status_val,
            "total": total_rows,
            "success": success_rows,
            "failed": failed_rows,
            "id": batch_id,
        })

        # Fetch final batch state
        fetch_query = text("""
            SELECT id, batch_no, import_type, file_name, status, total_rows, success_rows, failed_rows, started_at, finished_at, created_at
            FROM import_batches WHERE id = :id
        """)
        result = await session.execute(fetch_query, {"id": batch_id})
        row = result.fetchone()

        return ImportBatchResponse(
            id=row[0],
            batch_no=row[1],
            import_type=row[2],
            file_name=row[3],
            status=row[4],
            total_rows=row[5],
            success_rows=row[6],
            failed_rows=row[7],
            started_at=row[8].isoformat() if row[8] else "",
            finished_at=row[9].isoformat() if row[9] else None,
            created_at=row[10].isoformat() if row[10] else "",
        )

    except Exception as e:
        # Mark batch as failed
        update_query = text("""
            UPDATE import_batches
            SET status = 'failed', finished_at = NOW(), updated_at = NOW()
            WHERE id = :id
        """)
        await session.execute(update_query, {"id": batch_id})

        raise ValueError(f"导入失败: {str(e)}")


@router.post("/attendance", response_model=ImportBatchResponse, status_code=status.HTTP_201_CREATED)
async def import_course_attendance(
    current_user: CurrentUser,
    session: AsyncSession = Depends(get_session),
    file: UploadFile = File(...),
):
    """Import course attendance records from Excel/CSV.

    Required columns: customer_phone, course_name, class_date, status
    status should be one of: attended, absent, leave
    """
    if not is_admin_or_above(current_user.role_name):
        raise ForbiddenException("Admin access required")

    # Generate batch number
    batch_no = f"CA{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:6].upper()}"

    # Create batch record
    insert_query = text("""
        INSERT INTO import_batches (batch_no, import_type, file_name, status, total_rows, success_rows, failed_rows, started_at, created_by, created_at, updated_at)
        VALUES (:batch_no, 'course_attendance', :file_name, 'processing', 0, 0, 0, NOW(), :created_by, NOW(), NOW())
        RETURNING id
    """)
    result = await session.execute(insert_query, {
        "batch_no": batch_no,
        "file_name": file.filename,
        "created_by": current_user.id,
    })
    batch_id = result.fetchone()[0]

    try:
        # Parse file
        rows = await _parse_excel_or_csv(file)
        total_rows = len(rows)
        success_rows = 0
        failed_rows = 0

        valid_statuses = ['attended', 'absent', 'leave']

        for idx, row_data in enumerate(rows, start=2):
            try:
                # Validate required fields
                phone = str(row_data.get('customer_phone', '')).strip()
                course_name = str(row_data.get('course_name', '')).strip()
                class_date = row_data.get('class_date')
                status_val = str(row_data.get('status', '')).strip().lower()

                if not phone:
                    await _record_import_error(session, batch_id, idx, "MISSING_PHONE", "缺少客户手机号", dict(row_data))
                    failed_rows += 1
                    continue

                if not course_name:
                    await _record_import_error(session, batch_id, idx, "MISSING_COURSE", "缺少课程名称", dict(row_data))
                    failed_rows += 1
                    continue

                if not class_date:
                    await _record_import_error(session, batch_id, idx, "MISSING_DATE", "缺少上课日期", dict(row_data))
                    failed_rows += 1
                    continue

                if status_val not in valid_statuses:
                    await _record_import_error(session, batch_id, idx, "INVALID_STATUS", f"无效的状态值: {status_val}，有效值为: attended, absent, leave", dict(row_data))
                    failed_rows += 1
                    continue

                # Find customer
                customer_id = await _find_customer_by_phone(session, phone)
                if not customer_id:
                    await _record_import_error(session, batch_id, idx, "CUSTOMER_NOT_FOUND", f"找不到手机号为 {phone} 的客户", dict(row_data))
                    failed_rows += 1
                    continue

                # Parse date
                if isinstance(class_date, str):
                    from dateutil import parser
                    class_date = parser.parse(class_date).date()
                elif hasattr(class_date, 'date'):
                    class_date = class_date.date()

                # Insert record
                insert_record = text("""
                    INSERT INTO course_attendance_records
                    (customer_id, course_name, class_date, status, import_batch_id, created_at, updated_at)
                    VALUES (:customer_id, :course_name, :class_date, :status, :batch_id, NOW(), NOW())
                """)
                await session.execute(insert_record, {
                    "customer_id": customer_id,
                    "course_name": course_name,
                    "class_date": class_date,
                    "status": status_val,
                    "batch_id": batch_id,
                })
                success_rows += 1

            except Exception as e:
                await _record_import_error(session, batch_id, idx, "PARSE_ERROR", str(e), dict(row_data))
                failed_rows += 1

        # Update batch status
        final_status = "completed" if failed_rows == 0 else ("partial_success" if success_rows > 0 else "failed")
        update_query = text("""
            UPDATE import_batches
            SET status = :status, total_rows = :total, success_rows = :success, failed_rows = :failed, finished_at = NOW(), updated_at = NOW()
            WHERE id = :id
        """)
        await session.execute(update_query, {
            "status": final_status,
            "total": total_rows,
            "success": success_rows,
            "failed": failed_rows,
            "id": batch_id,
        })

        # Fetch final batch state
        fetch_query = text("""
            SELECT id, batch_no, import_type, file_name, status, total_rows, success_rows, failed_rows, started_at, finished_at, created_at
            FROM import_batches WHERE id = :id
        """)
        result = await session.execute(fetch_query, {"id": batch_id})
        row = result.fetchone()

        return ImportBatchResponse(
            id=row[0],
            batch_no=row[1],
            import_type=row[2],
            file_name=row[3],
            status=row[4],
            total_rows=row[5],
            success_rows=row[6],
            failed_rows=row[7],
            started_at=row[8].isoformat() if row[8] else "",
            finished_at=row[9].isoformat() if row[9] else None,
            created_at=row[10].isoformat() if row[10] else "",
        )

    except Exception as e:
        # Mark batch as failed
        update_query = text("""
            UPDATE import_batches
            SET status = 'failed', finished_at = NOW(), updated_at = NOW()
            WHERE id = :id
        """)
        await session.execute(update_query, {"id": batch_id})

        raise ValueError(f"导入失败: {str(e)}")
